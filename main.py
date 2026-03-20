import pandas as pd
import re
import asyncio
from datetime import datetime
from zoneinfo import ZoneInfo
from playwright.async_api import async_playwright
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side

INPUT_FILE = "tweets_input.xlsx"
OUTPUT_FILE = "scraped_tweets_output.xlsx"
MAX_CONCURRENT = 5


async def extract_views(page):
    try:
        spans = await page.locator("span").all()
        for span in spans:
            text = await span.inner_text()
            text = text.strip()
            if text.lower().endswith("views"):
                number = re.sub(r"[^\d]", "", text)
                if number:
                    return int(number)

        analytics_link = page.locator('a[aria-label*="View post analytics"]').first
        if await analytics_link.count() > 0:
            aria_label = await analytics_link.get_attribute("aria-label")
            if aria_label:
                match = re.search(r'(\d[\d,\.]*)\s*[Vv]iews?', aria_label)
                if match:
                    number = re.sub(r"[^\d]", "", match.group(1))
                    if number:
                        return int(number)

        all_text = await page.content()
        matches = re.findall(r'(\d[\d,\.]+)\s*[Vv]iews?', all_text)
        if matches:
            for match in matches:
                number = re.sub(r"[^\d]", "", match)
                if number and int(number) > 0:
                    return int(number)

    except Exception as e:
        pass
    return None


async def extract_likes(page):
    try:
        like_btn = page.locator('[data-testid="like"]').first
        if await like_btn.count() > 0:
            aria_label = await like_btn.get_attribute("aria-label")
            if aria_label:
                match = re.search(r'(\d[\d,\.]*)', aria_label)
                if match:
                    number = re.sub(r"[^\d]", "", match.group(1))
                    if number:
                        return int(number)

        like_spans = await page.locator('[data-testid="like"] span').all()
        for span in like_spans:
            text = (await span.inner_text()).strip()
            if re.match(r'^\d[\d,\.]*[KkMm]?$', text):
                return parse_count(text)

        unlike_btn = page.locator('[data-testid="unlike"]').first
        if await unlike_btn.count() > 0:
            aria_label = await unlike_btn.get_attribute("aria-label")
            if aria_label:
                match = re.search(r'(\d[\d,\.]*)', aria_label)
                if match:
                    number = re.sub(r"[^\d]", "", match.group(1))
                    if number:
                        return int(number)

    except Exception as e:
        pass
    return None


async def extract_retweets(page):
    try:
        rt_btn = page.locator('[data-testid="retweet"]').first
        if await rt_btn.count() > 0:
            aria_label = await rt_btn.get_attribute("aria-label")
            if aria_label:
                match = re.search(r'(\d[\d,\.]*)', aria_label)
                if match:
                    number = re.sub(r"[^\d]", "", match.group(1))
                    if number:
                        return int(number)

        rt_spans = await page.locator('[data-testid="retweet"] span').all()
        for span in rt_spans:
            text = (await span.inner_text()).strip()
            if re.match(r'^\d[\d,\.]*[KkMm]?$', text):
                return parse_count(text)

        unrt_btn = page.locator('[data-testid="unretweet"]').first
        if await unrt_btn.count() > 0:
            aria_label = await unrt_btn.get_attribute("aria-label")
            if aria_label:
                match = re.search(r'(\d[\d,\.]*)', aria_label)
                if match:
                    number = re.sub(r"[^\d]", "", match.group(1))
                    if number:
                        return int(number)

    except Exception as e:
        pass
    return None


async def extract_comments(page):
    try:
        reply_btn = page.locator('[data-testid="reply"]').first
        if await reply_btn.count() > 0:
            aria_label = await reply_btn.get_attribute("aria-label")
            if aria_label:
                match = re.search(r'(\d[\d,\.]*)', aria_label)
                if match:
                    number = re.sub(r"[^\d]", "", match.group(1))
                    if number:
                        return int(number)

        reply_spans = await page.locator('[data-testid="reply"] span').all()
        for span in reply_spans:
            text = (await span.inner_text()).strip()
            if re.match(r'^\d[\d,\.]*[KkMm]?$', text):
                return parse_count(text)

    except Exception as e:
        pass
    return None


def parse_count(text):
    text = text.strip().replace(",", "")
    try:
        if text.lower().endswith("k"):
            return int(float(text[:-1]) * 1_000)
        elif text.lower().endswith("m"):
            return int(float(text[:-1]) * 1_000_000)
        else:
            return int(re.sub(r"[^\d]", "", text))
    except:
        return None


async def extract_datetime(page, url):
    try:
        status_match = re.search(r'/status/(\d+)', url)
        if not status_match:
            return None

        status_id = status_match.group(1)

        time_elements = await page.locator("time").all()

        for time_el in time_elements:
            dt_raw = await time_el.get_attribute("datetime")
            if not dt_raw:
                continue

            parent_link = time_el.locator("xpath=ancestor::a[contains(@href, '/status/')]").first
            if await parent_link.count() > 0:
                href = await parent_link.get_attribute("href")

                if status_id in href:
                    dt_utc = datetime.fromisoformat(
                        dt_raw.replace("Z", "+00:00")
                    ).astimezone(ZoneInfo("Europe/Belgrade"))
                    return dt_utc.strftime("%d/%m/%Y %H:%M")

        meta_time = await page.locator('meta[property="article:published_time"]').first.get_attribute("content")
        if meta_time:
            dt_utc = datetime.fromisoformat(
                meta_time.replace("Z", "+00:00")
            ).astimezone(ZoneInfo("Europe/Belgrade"))
            return dt_utc.strftime("%d/%m/%Y %H:%M")

    except Exception as e:
        pass
    return None


def extract_account_link(url):
    try:
        match = re.search(r"x\.com/([^/]+)/status", url)
        if match:
            return f"https://x.com/{match.group(1)}"
    except:
        pass
    return None


async def process_url(browser, index, url, df, total):
    page = await browser.new_page()
    try:
        print(f"{index + 1}/{total} -> {url}")

        if not isinstance(url, str) or "x.com" not in url:
            return

        await page.goto(url, timeout=30000, wait_until="domcontentloaded")
        await page.wait_for_timeout(10000)

        final_url = page.url

        df.at[index, "reach"] = await extract_views(page)
        df.at[index, "likes"] = await extract_likes(page)
        df.at[index, "retweets"] = await extract_retweets(page)
        df.at[index, "comments"] = await extract_comments(page)
        df.at[index, "datetime"] = await extract_datetime(page, final_url)
        df.at[index, "account_link"] = extract_account_link(final_url)

        await asyncio.sleep(2)

    except Exception as e:
        pass
    finally:
        await page.close()


def format_excel(file_path):
    wb = load_workbook(file_path)
    ws = wb.active

    arial_font = Font(name='Arial', size=10)
    center_alignment = Alignment(horizontal='center', vertical='center')
    right_alignment = Alignment(horizontal='right', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.font = arial_font
            cell.alignment = center_alignment
            cell.border = thin_border

    for col_idx in range(1, ws.max_column + 1):
        header = ws.cell(row=1, column=col_idx).value
        if header in ("reach", "likes", "retweets", "comments"):
            col_letter = ws.cell(row=1, column=col_idx).column_letter
            for row in range(1, ws.max_row + 1):
                ws[f"{col_letter}{row}"].alignment = right_alignment

    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter

        for cell in column:
            try:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            except:
                pass

        adjusted_width = min(max_length + 2, 100)
        ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(file_path)


async def main():
    df = pd.read_excel(INPUT_FILE, header=None)

    df.rename(columns={0: "url"}, inplace=True)

    df["datetime"] = None
    df["account_link"] = None
    df["reach"] = None
    df["likes"] = None
    df["retweets"] = None
    df["comments"] = None

    df = df[["datetime", "account_link", "url", "reach", "likes", "retweets", "comments"]]

    total_count = len(df["url"])

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)

        tasks = []
        for i, row in df.iterrows():
            url = row["url"]
            tasks.append(process_url(browser, i, url, df, total_count))

        for i in range(0, len(tasks), MAX_CONCURRENT):
            batch = tasks[i:i + MAX_CONCURRENT]
            await asyncio.gather(*batch)

        await browser.close()

    df.to_excel(OUTPUT_FILE, index=False)
    format_excel(OUTPUT_FILE)

    print(f"Finished! Results saved in: {OUTPUT_FILE}")


if __name__ == "__main__":
    asyncio.run(main())